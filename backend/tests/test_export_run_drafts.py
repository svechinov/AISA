"""Фаза 2, Task 4: XLSX-экспорт черновиков рана (экспортный канал FG, решение B владельца 02.09).

Движок для этого канала не отправляет: менеджеры FG рассылают со своих ящиков, читая этот файл.
Отсюда состав колонок и пустая колонка «Менеджер» (правило разбивки базы — вопрос Г6 к FG).
0 tokens: ни LLM, ни сети."""

from __future__ import annotations

import sys
from pathlib import Path

SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import export_run_drafts as export  # noqa: E402
from app.models.contact import Contact  # noqa: E402
from app.models.email_draft import EmailDraft  # noqa: E402
from app.repositories.project_repo import create_project  # noqa: E402
from app.repositories.run_repo import create_run  # noqa: E402

BODY = (
    "Здравствуйте, Мария!\n\n"
    "Первый абзац письма.\n\n"
    "Второй абзац письма."
)


def _run_with_drafts(fresh_db, count: int):
    proj = create_project(fresh_db, name="ExportTest", type="generic")
    run = create_run(fresh_db, project_id=proj.id, workflow_name="generic_outreach", input_json={})
    for i in range(count):
        contact = Contact(run_id=run.id, name=f"Контакт {i}", role="Директор по персоналу",
                          email=f"c{i}@example.com", company=f"Компания {i}", source_json={})
        fresh_db.add(contact)
        fresh_db.flush()
        fresh_db.add(EmailDraft(
            run_id=run.id, contact_id=contact.id, company=contact.company,
            to_email=contact.email, subject=f"Тема {i}", body=BODY,
        ))
    fresh_db.commit()
    return run


def test_collect_rows_returns_one_row_per_draft(fresh_db):
    run = _run_with_drafts(fresh_db, 3)
    rows = export.collect_rows(fresh_db, run.id)
    assert len(rows) == 3
    assert [r["company"] for r in rows] == ["Компания 0", "Компания 1", "Компания 2"]
    assert rows[0]["contact"] == "Контакт 0"
    assert rows[0]["email"] == "c0@example.com"
    assert rows[0]["role"] == "Директор по персоналу"
    assert rows[0]["subject"] == "Тема 0"
    assert rows[0]["body"] == BODY
    assert rows[0]["manager"] == ""  # правило разбивки ждёт ответа FG на Г6


def test_collect_rows_survives_a_missing_contact(fresh_db):
    """Черновик денормализует company/to_email — строка выгружается даже без строки контакта."""
    run = _run_with_drafts(fresh_db, 1)
    draft = fresh_db.query(EmailDraft).filter(EmailDraft.run_id == run.id).one()
    fresh_db.delete(fresh_db.get(Contact, draft.contact_id))
    fresh_db.commit()

    rows = export.collect_rows(fresh_db, run.id)
    assert len(rows) == 1
    assert rows[0]["company"] == "Компания 0"
    assert rows[0]["contact"] == ""
    assert rows[0]["role"] == ""


def test_write_xlsx_opens_and_has_header_and_all_rows(fresh_db, tmp_path):
    import openpyxl

    run = _run_with_drafts(fresh_db, 4)
    rows = export.collect_rows(fresh_db, run.id)
    out = tmp_path / "run.xlsx"

    export.write_xlsx(rows, out)

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert [c.value for c in ws[1]] == export.HEADER
    assert ws.max_row == len(rows) + 1  # шапка + по строке на черновик
    assert ws.cell(row=2, column=export.HEADER.index("Компания") + 1).value == "Компания 0"
    body_cell = ws.cell(row=2, column=export.HEADER.index("Тело письма") + 1)
    assert body_cell.value == BODY
    assert body_cell.alignment.wrap_text is True


def test_empty_run_writes_header_only(fresh_db, tmp_path):
    import openpyxl

    proj = create_project(fresh_db, name="ExportEmpty", type="generic")
    run = create_run(fresh_db, project_id=proj.id, workflow_name="generic_outreach", input_json={})
    fresh_db.commit()
    out = tmp_path / "empty.xlsx"

    export.write_xlsx(export.collect_rows(fresh_db, run.id), out)

    ws = openpyxl.load_workbook(out).active
    assert ws.max_row == 1

"""Выгрузка черновиков рана в XLSX для ручной рассылки менеджерами (Фаза 2, Task 4).

Экспортный канал FG (решение B владельца 02.09.2026): движок ГЕНЕРИРУЕТ письма, а отправляют их
менеджеры FG вручную со своих ящиков по существующей базе клиента. Поэтому:

- черновики НЕ аппрувятся (approve в UI = автопостановка в очередь отправки) — выгружаются как
  есть, в том состоянии, в каком их оставило ревью;
- в теле письма нет подписи: она клеится при отправке из run_setups.sender_signature_html, а у
  FG-ранов это поле пустое (что заодно физически блокирует отправку —
  email_sender.validate_outbound_draft_sendable);
- колонка «Менеджер» выгружается пустой: правило разбивки базы между менеджерами — открытый
  вопрос Г6 к FG. Заполняется вручную в файле.

XLSX (не CSV): адресат — менеджеры, им нужен файл, который открывается двойным кликом, с
переносом строк в теле письма. openpyxl напрямую, без pandas: нужен контроль ширины колонок и
wrap_text.

Usage:
    cd backend && venv/Scripts/python.exe scripts/export_run_drafts.py --run-id 7
    cd backend && venv/Scripts/python.exe scripts/export_run_drafts.py --run-id 7 --out C:/tmp/fg.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.init_db import ensure_schema  # noqa: E402
from app.models.contact import Contact  # noqa: E402
from app.models.email_draft import EmailDraft  # noqa: E402
from app.models.run import Run  # noqa: E402

HEADER: list[str] = [
    "ID черновика", "Компания", "Контакт", "Должность", "Email", "Тема", "Тело письма", "Менеджер",
]

# Ширины под HEADER; тело письма — широкая колонка с переносом.
_COLUMN_WIDTHS: tuple[int, ...] = (12, 28, 22, 24, 30, 44, 90, 18)
_BODY_COLUMN = HEADER.index("Тело письма") + 1


def collect_rows(db, run_id: int) -> list[dict]:
    """Черновики рана в порядке id. company/to_email берём из самого черновика (он их
    денормализует при создании — email_draft_repo.create_email_draft), должность и имя — из
    контакта, если строка контакта ещё жива."""
    drafts = (
        db.query(EmailDraft)
        .filter(EmailDraft.run_id == run_id)
        .order_by(EmailDraft.id.asc())
        .all()
    )
    rows: list[dict] = []
    for draft in drafts:
        contact = db.get(Contact, draft.contact_id) if draft.contact_id else None
        rows.append(
            {
                "draft_id": draft.id,
                "company": (draft.company or getattr(contact, "company", None) or "").strip(),
                "contact": (getattr(contact, "name", None) or "").strip(),
                "role": (getattr(contact, "role", None) or "").strip(),
                "email": (draft.to_email or getattr(contact, "email", None) or "").strip(),
                "subject": (draft.subject or "").strip(),
                "body": (draft.body or "").strip(),
                "manager": "",
            }
        )
    return rows


def write_xlsx(rows: list[dict], path: Path) -> Path:
    """Один лист: шапка + по строке на черновик. Возвращает путь записанного файла."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Черновики"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            row["draft_id"], row["company"], row["contact"], row["role"],
            row["email"], row["subject"], row["body"], row["manager"],
        ])

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    for excel_row in range(2, len(rows) + 2):
        ws.cell(row=excel_row, column=_BODY_COLUMN).alignment = Alignment(
            wrap_text=True, vertical="top",
        )

    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def main() -> None:
    ap = argparse.ArgumentParser(description="Выгрузить черновики рана в XLSX для ручной рассылки.")
    ap.add_argument("--run-id", type=int, required=True, help="Ран, чьи черновики выгружаем.")
    ap.add_argument(
        "--out", type=str, default=None,
        help="Путь к файлу. По умолчанию backend/exports/run_<id>_drafts.xlsx.",
    )
    args = ap.parse_args()

    ensure_schema()
    db = SessionLocal()
    try:
        run = db.query(Run).filter(Run.id == args.run_id).first()
        if not run:
            print(f"Run id={args.run_id} not found.", file=sys.stderr)
            sys.exit(1)

        rows = collect_rows(db, run.id)
        out = Path(args.out) if args.out else (
            Path(__file__).resolve().parents[1] / "exports" / f"run_{run.id}_drafts.xlsx"
        )
        write_xlsx(rows, out)
        print(f"Exported {len(rows)} draft(s) of run_id={run.id} ({run.name!r}) -> {out}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
